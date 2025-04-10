import tkinter as tk
from tkinter import filedialog, messagebox
from PIL import Image, ImageTk
import openpyxl
import os

class ImageTextEditorApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Image-Text Editor")

        self.image_label = tk.Label(root)
        self.image_label.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        self.text_frame = tk.Frame(root)
        self.text_frame.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True)

        self.text_widget = tk.Text(self.text_frame, wrap=tk.WORD)
        self.text_widget.pack(fill=tk.BOTH, expand=True)

        self.next_button = tk.Button(self.text_frame, text="Next", command=self.next_entry)
        self.next_button.pack(pady=10)

        self.data = []
        self.current_index = 0
        self.wb = None
        self.ws = None

        self.load_file()

    def load_file(self):
        filepath = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
        if not filepath:
            self.root.quit()
            return

        self.wb = openpyxl.load_workbook(filepath)
        self.ws = self.wb.active
        self.filepath = filepath

        self.data = list(self.ws.iter_rows(min_row=2, values_only=False))
        self.show_entry()

    def show_entry(self):
        if self.current_index >= len(self.data):
            messagebox.showinfo("Done", "No more entries.")
            return

        row = self.data[self.current_index]
        image_cell = row[0]
        text_cell = row[1]
        self.flag_cell = row[2] if len(row) > 2 else self.ws.cell(row=image_cell.row, column=3)

        image_path = image_cell.value
        text = text_cell.value if text_cell.value else ""

        if not os.path.exists(image_path):
            messagebox.showerror("Error", f"Image not found: {image_path}")
            return

        img = Image.open(image_path)
        window_width = self.root.winfo_screenwidth()
        window_height = self.root.winfo_screenheight()

        max_width = window_width // 2
        max_height = window_height

        img.thumbnail((max_width, max_height))
        self.tk_img = ImageTk.PhotoImage(img)

        self.image_label.config(image=self.tk_img)
        self.text_widget.delete("1.0", tk.END)
        self.text_widget.insert(tk.END, text)

    def next_entry(self):
        if self.current_index >= len(self.data):
            return

        text = self.text_widget.get("1.0", tk.END).strip()
        text_cell = self.data[self.current_index][1]
        text_cell.value = text
        self.flag_cell.value = "edited"

        self.wb.save(self.filepath)

        self.current_index += 1
        self.show_entry()

if __name__ == "__main__":
    root = tk.Tk()
    app = ImageTextEditorApp(root)
    root.mainloop()
